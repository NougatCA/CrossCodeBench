from dataclasses import dataclass
import os
import json
from tqdm import tqdm
from typing import Union
import re


@dataclass
class DataInstance:
    inputs: Union[str, list[str]]
    outputs: Union[str, list[str]]
    split: str
    idx: str


def walk_tasks(task_dir):
    assert os.path.isdir(task_dir)
    for file_name in os.listdir(task_dir):
        file_path = os.path.join(task_dir, file_name)
        if file_name.endswith(".meta.json") and os.path.isfile(file_path):
            task_name = re.sub(r".meta.json$", "", file_name)
            data_file_name = f"{task_name}.data.json"
            data_file_path = os.path.join(task_dir, data_file_name)
            assert os.path.isfile(data_file_path), f"The data file of task `{task_name}` not found."
            yield file_name


def read_devign(data_dir):
    data_dir = os.path.join(data_dir, "devign")

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"{split}.jsonl"), mode="r", encoding="utf-8") as f:
            for line in f.readlines():
                js = json.loads(line.strip())
                label = "Yes" if js["target"] == 1 else "No"
                instances.append(
                    DataInstance(
                        inputs=js["func"],
                        outputs=label,
                        split=split,
                        idx=js["idx"]
                    )
                )
                sizes[split] += 1
    sizes["total"] = len(instances)
    return instances, sizes


def read_bigclonebench(data_dir):
    data_dir = os.path.join(data_dir, "bigclonebench")

    aux_data = {}
    with open(os.path.join(data_dir, "data.jsonl"), mode="r", encoding="utf-8") as f:
        for line in f.readlines():
            js = json.loads(line.strip())
            code = ' '.join(js['func'].split())
            aux_data[js['idx']] = code

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"{split}.txt"), mode="r", encoding="utf-8") as f:
            lines = f.readlines()
        for idx, line in enumerate(tqdm(lines, desc="Reading", total=len(lines))):
            url1, url2, label = line.strip().split('\t')
            if url1 not in aux_data or url2 not in aux_data:
                continue
            label = "Yes" if label.strip() == "1" else "No"
            instances.append(
                DataInstance(
                    inputs=[aux_data[url1], aux_data[url2]],
                    outputs=label,
                    split=split,
                    idx=str(idx)
                )
            )
            sizes[split] += 1
    sizes["total"] = len(instances)
    return instances, sizes


def read_exception_type(data_dir):
    data_dir = os.path.join(data_dir, "exception_type")

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"{split}.jsonl"), mode="r", encoding="utf-8") as f:
            lines = f.readlines()
        for idx, line in enumerate(tqdm(lines, desc="Reading", total=len(lines))):
            js = json.loads(line.strip())
            code = js["function"]
            target_txt = js["label"]
            instances.append(
                DataInstance(
                    inputs=code,
                    outputs=target_txt,
                    split=split,
                    idx=str(idx)
                )
            )
            sizes[split] += 1
    sizes["total"] = len(instances)
    return instances, sizes


def read_function_docstring_mismatch(data_dir):
    data_dir = os.path.join(data_dir, "function_docstring_mismatch")

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"{split}.jsonl"), mode="r", encoding="utf-8") as f:
            lines = f.readlines()
        for idx, line in enumerate(tqdm(lines, desc="Reading", total=len(lines))):
            js = json.loads(line.strip())
            code = js["function"]
            doc = js["docstring"]
            target_txt = js["label"]
            instances.append(
                DataInstance(
                    inputs=[code, doc],
                    outputs=target_txt,
                    split=split,
                    idx=str(idx)
                )
            )
            sizes[split] += 1
    sizes["total"] = len(instances)
    return instances, sizes


def read_variable_misuse(data_dir):
    data_dir = os.path.join(data_dir, "variable_misuse")

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"{split}.jsonl"), mode="r", encoding="utf-8") as f:
            lines = f.readlines()
        for idx, line in enumerate(tqdm(lines, desc="Reading", total=len(lines))):
            js = json.loads(line.strip())
            code = js["function"]
            target_txt = js["label"]
            instances.append(
                DataInstance(
                    inputs=code,
                    outputs=target_txt,
                    split=split,
                    idx=str(idx)
                )
            )
            sizes[split] += 1
    sizes["total"] = len(instances)
    return instances, sizes


def read_swapped_operands(data_dir):
    data_dir = os.path.join(data_dir, "swapped_operands")

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"{split}.jsonl"), mode="r", encoding="utf-8") as f:
            lines = f.readlines()
        for idx, line in enumerate(tqdm(lines, desc="Reading", total=len(lines))):
            js = json.loads(line.strip())
            code = js["function"]
            target_txt = js["label"]
            instances.append(
                DataInstance(
                    inputs=code,
                    outputs=target_txt,
                    split=split,
                    idx=str(idx)
                )
            )
            sizes[split] += 1
    sizes["total"] = len(instances)
    return instances, sizes


def read_wrong_binary_operator(data_dir):
    data_dir = os.path.join(data_dir, "wrong_binary_operator")

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"{split}.jsonl"), mode="r", encoding="utf-8") as f:
            lines = f.readlines()
        for idx, line in enumerate(tqdm(lines, desc="Reading", total=len(lines))):
            js = json.loads(line.strip())
            code = js["function"]
            target_txt = js["label"]
            instances.append(
                DataInstance(
                    inputs=code,
                    outputs=target_txt,
                    split=split,
                    idx=str(idx)
                )
            )
            sizes[split] += 1
    sizes["total"] = len(instances)
    return instances, sizes


def read_web_query_test(data_dir):
    data_dir = os.path.join(data_dir, "web_query_test")

    instances = []
    sizes = {
        "test": 0
    }

    idx_to_answer = {}
    with open(os.path.join(data_dir, "answers.txt"), mode="r", encoding="utf-8") as f:
        for line in f:
            idx, answer = line.strip().split("\t")
            idx_to_answer[idx] = int(answer)

    with open(os.path.join(data_dir, "test_webquery.json")) as f:
        data = json.load(f)
    for js in tqdm(data, total=len(data), desc=f"Reading"):
        code = js["code"]
        nl = js["doc"]
        idx = js["idx"]
        label = "Yes" if idx_to_answer[idx] == 1 else "No"
        instances.append(
            DataInstance(
                inputs=[nl, code],
                outputs=label,
                split="test",
                idx=idx
            )
        )
        sizes["test"] += 1
    sizes["total"] = len(instances)
    return instances, sizes


def read_cosqa(data_dir):
    data_dir = os.path.join(data_dir, "cosqa")

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
    }

    for split in ["train", "valid"]:
        with open(os.path.join(data_dir, f"cosqa-{split}.json"), mode="r", encoding="utf-8") as f:
            data = json.load(f)
        for js in tqdm(data, total=len(data), desc=f"Reading"):
            code = js["code"]
            nl = js["doc"]
            idx = js["idx"]
            label = "Yes" if js["label"] == 1 else "No"
            instances.append(
                DataInstance(
                    inputs=[nl, code],
                    outputs=label,
                    split=split,
                    idx=idx
                )
            )
            sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_code_trans_java_cs(data_dir):
    data_dir = os.path.join(data_dir, "code_trans")

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"{split}.java-cs.txt.java"), mode="r", encoding="utf-8") as src_f, \
                open(os.path.join(data_dir, f"{split}.java-cs.txt.cs"), mode="r", encoding="utf-8") as tgt_f:
            sources = src_f.readlines()
            targets = tgt_f.readlines()
        assert len(sources) == len(targets)
        for idx, (source, target) in enumerate(tqdm(zip(sources, targets), desc="Reading", total=len(sources))):
            instances.append(
                DataInstance(
                    inputs=source.strip(),
                    outputs=target.strip(),
                    split=split,
                    idx=str(idx)
                )
            )
            sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_code_trans_cs_java(data_dir):
    data_dir = os.path.join(data_dir, "code_trans")

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"{split}.java-cs.txt.cs"), mode="r", encoding="utf-8") as src_f, \
                open(os.path.join(data_dir, f"{split}.java-cs.txt.java"), mode="r", encoding="utf-8") as tgt_f:
            sources = src_f.readlines()
            targets = tgt_f.readlines()
        assert len(sources) == len(targets)
        for idx, (source, target) in enumerate(tqdm(zip(sources, targets), desc="Reading", total=len(sources))):
            instances.append(
                DataInstance(
                    inputs=source.strip(),
                    outputs=target.strip(),
                    split=split,
                    idx=str(idx)
                )
            )
            sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_trans_coder(data_dir, source_lang, target_lang):
    assert source_lang in ["java", "python", "cpp"]
    assert target_lang in ["java", "python", "cpp"]

    data_dir = os.path.join(data_dir, "trans_coder")
    instances = []
    sizes = {
        "valid": 0,
        "test": 0
    }
    for split in ["valid", "test"]:
        with open(os.path.join(data_dir, f"transcoder_{split}.{source_lang}.tok"), mode="r", encoding="utf-8") as src_f, \
                open(os.path.join(data_dir, f"transcoder_{split}.{target_lang}.tok"), mode="r",
                     encoding="utf-8") as tgt_f:
            sources = src_f.readlines()
            targets = tgt_f.readlines()
        assert len(sources) == len(targets)
        for idx, (source, target) in enumerate(tqdm(zip(sources, targets), desc="Reading", total=len(sources))):
            src_idx, source = source.split(" | ", 1)
            src_idx = src_idx.strip()
            source = source.strip()

            tgt_idx, target = target.split(" | ", 1)
            tgt_idx = tgt_idx.strip()
            target = target.strip()

            assert src_idx == tgt_idx

            instances.append(
                DataInstance(
                    inputs=source,
                    outputs=target,
                    split=split,
                    idx=src_idx
                )
            )
            sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_bfp(data_dir, subset):
    assert subset in ["small", "medium"]
    data_dir = os.path.join(data_dir, "bfp", subset)

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"{split}.buggy-fixed.buggy"), mode="r", encoding="utf-8") as src_f, \
                open(os.path.join(data_dir, f"{split}.buggy-fixed.fixed"), mode="r", encoding="utf-8") as tgt_f:
            sources = src_f.readlines()
            targets = tgt_f.readlines()
        assert len(sources) == len(targets)
        for idx, (source, target) in enumerate(tqdm(zip(sources, targets), desc="Reading", total=len(sources))):
            instances.append(
                DataInstance(
                    inputs=source.strip(),
                    outputs=target.strip(),
                    split=split,
                    idx=str(idx)
                )
            )
            sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_any_code_completion(data_dir):
    data_dir = os.path.join(data_dir, "any_code_completion")

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"data.TargetType.seq.{split}.source.txt"), mode="r",
                  encoding="utf-8") as src_f, \
                open(os.path.join(data_dir, f"data.TargetType.seq.{split}.target.txt"), mode="r",
                     encoding="utf-8") as tgt_f:
            sources = src_f.readlines()
            targets = tgt_f.readlines()
        assert len(sources) == len(targets)
        for idx, (source, target) in enumerate(tqdm(zip(sources, targets), desc="Reading", total=len(sources))):
            instances.append(
                DataInstance(
                    inputs=source.strip(),
                    outputs=target.strip(),
                    split=split,
                    idx=str(idx)
                )
            )
            sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_tp_mutant(data_dir):
    data_dir = os.path.join(data_dir, "tp_mutant")

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"{split}_fixed.txt"), mode="r", encoding="utf-8") as src_f, \
                open(os.path.join(data_dir, f"{split}_buggy.txt"), mode="r", encoding="utf-8") as tgt_f:
            sources = src_f.readlines()
            targets = tgt_f.readlines()
        assert len(sources) == len(targets)
        for idx, (source, target) in enumerate(tqdm(zip(sources, targets), desc="Reading", total=len(sources))):
            instances.append(
                DataInstance(
                    inputs=source.strip(),
                    outputs=target.strip(),
                    split=split,
                    idx=str(idx)
                )
            )
            sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_tp_fixing(data_dir):
    data_dir = os.path.join(data_dir, "tp_mutant")

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"{split}_buggy.txt"), mode="r", encoding="utf-8") as src_f, \
                open(os.path.join(data_dir, f"{split}_fixed.txt"), mode="r", encoding="utf-8") as tgt_f:
            sources = src_f.readlines()
            targets = tgt_f.readlines()
        assert len(sources) == len(targets)
        for idx, (source, target) in enumerate(tqdm(zip(sources, targets), desc="Reading", total=len(sources))):
            instances.append(
                DataInstance(
                    inputs=source.strip(),
                    outputs=target.strip(),
                    split=split,
                    idx=str(idx)
                )
            )
            sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_tap(data_dir, subset):
    assert subset in ["abs", "raw"]

    data_dir = os.path.join(data_dir, "tap", subset)

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"{split}_methods.txt"), mode="r", encoding="utf-8") as src_f, \
                open(os.path.join(data_dir, f"{split}_assert.txt"), mode="r", encoding="utf-8") as tgt_f:
            sources = src_f.readlines()
            targets = tgt_f.readlines()
        assert len(sources) == len(targets)
        for idx, (source, target) in enumerate(tqdm(zip(sources, targets), desc="Reading", total=len(sources))):
            instances.append(
                DataInstance(
                    inputs=source.strip(),
                    outputs=target.strip(),
                    split=split,
                    idx=str(idx)
                )
            )
            sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_code_search_net_filtered(data_dir, subset):
    assert subset in ["java", "python", "javascript", "php", "go", "ruby"]

    data_dir = os.path.join(data_dir, "code_search_net_filtered", subset)

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"{split}.jsonl"), mode="r", encoding="utf-8") as f:
            lines = f.readlines()
        for idx, line in enumerate(tqdm(lines, total=len(lines), desc=f"Loading {split} data")):
            js = json.loads(line.strip())
            source = " ".join(js["code_tokens"])
            target = " ".join(js["docstring_tokens"])
            instances.append(
                DataInstance(
                    inputs=source,
                    outputs=target,
                    split=split,
                    idx=js["path"]
                )
            )
            sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_funcom_tokenized(data_dir):
    data_dir = os.path.join(data_dir, "funcom", "tokenized")

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"functions.{split}"), mode="r", encoding="utf-8") as src_f, \
                open(os.path.join(data_dir, f"comments.{split}"), mode="r", encoding="utf-8") as tgt_f:
            sources = src_f.readlines()
            targets = tgt_f.readlines()
        assert len(sources) == len(targets)
        for source, target in tqdm(zip(sources, targets), desc="Reading", total=len(sources)):
            src_idx, source = source.split("\t", 1)
            tgt_idx, target = target.split("\t", 1)
            assert src_idx == tgt_idx
            instances.append(
                DataInstance(
                    inputs=source.strip(),
                    outputs=target.strip(),
                    split=split,
                    idx=src_idx
                )
            )
            sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_funcom_raw(data_dir):
    data_dir = os.path.join(data_dir, "funcom", "raw")
    instances = []
    sizes = {
        "total": 0
    }
    with open(os.path.join(data_dir, "functions.json"), mode="r", encoding="utf-8") as f:
        fun_data = json.load(f)
    with open(os.path.join(data_dir, "comments.json"), mode="r", encoding="utf-8") as f:
        com_data = json.load(f)

    for k, v in fun_data.items():
        assert k in com_data
        function = v.strip()
        comment = com_data[k].strip()
        instances.append(
            DataInstance(
                inputs=function,
                outputs=comment,
                split="",
                idx=str(k)
            )
        )
        sizes["total"] += 1

    return instances, sizes


def read_deep_com(data_dir):
    data_dir = os.path.join(data_dir, "deep_com")

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }

    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"{split}.jsonl"), mode="r", encoding="utf-8") as f:
            lines = f.readlines()
        for idx, line in enumerate(tqdm(lines, desc="Reading", total=len(lines))):
            js = json.loads(line.strip())
            code = js["code"].strip()
            doc = js["nl"].strip()
            instances.append(
                DataInstance(
                    inputs=code,
                    outputs=doc,
                    split=split,
                    idx=str(idx)
                )
            )
            sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_tl_code_sum(data_dir):
    data_dir = os.path.join(data_dir, "tl_code_sum")

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }

    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"{split}.jsonl"), mode="r", encoding="utf-8") as f:
            lines = f.readlines()
        for idx, line in enumerate(tqdm(lines, desc="Reading", total=len(lines))):
            js = json.loads(line.strip())
            code = js["code"].strip()
            doc = js["comment"].strip()
            instances.append(
                DataInstance(
                    inputs=code,
                    outputs=doc,
                    split=split,
                    idx=str(js["id"])
                )
            )
            sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_tl_api_seq_sum(data_dir):
    data_dir = os.path.join(data_dir, "tl_code_sum")

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }

    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"{split}.jsonl"), mode="r", encoding="utf-8") as f:
            lines = f.readlines()
        for idx, line in enumerate(tqdm(lines, desc="Reading", total=len(lines))):
            js = json.loads(line.strip())
            api_seq = js["api_seq"]
            doc = js["comment"].strip()
            instances.append(
                DataInstance(
                    inputs=api_seq,
                    outputs=doc,
                    split=split,
                    idx=str(js["id"])
                )
            )
            sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_tl_api_seq_sum_large(data_dir):
    data_dir = os.path.join(data_dir, "api_sum")

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }

    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"{split}.jsonl"), mode="r", encoding="utf-8") as f:
            lines = f.readlines()
        for idx, line in enumerate(tqdm(lines, desc="Reading", total=len(lines))):
            js = json.loads(line.strip())
            api_seq = js["api_seq"]
            doc = js["comment"].strip()
            instances.append(
                DataInstance(
                    inputs=api_seq,
                    outputs=doc,
                    split=split,
                    idx=str(js["id"])
                )
            )
            sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_code_docstring_corpus_sum(data_dir):
    def escape_code(s):
        s = s.strip()
        # s = s.replace(" DCNL DCSP ", "\n")
        s = s.replace(" DCNL DCSP ", "\n\t")
        s = s.replace(" DCNL  DCSP ", "\n\t")
        s = s.replace(" DCNL ", "\n")
        s = s.replace(" DCSP ", " ")
        while "\t " in s:
            s = s.replace("\t ", "\t\t")
        return s

    def escape_nl(s):
        s = s.strip("'")
        s = s.strip()
        if "DCNL" in s:
            s = s.split("DCNL")[0]
        s = s.replace("DCSP", " ")
        return " ".join(s.split())

    data_dir = os.path.join(data_dir, "code_docstring_corpus")

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }

    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"data_ps.declbodies.{split}"), mode="r", encoding="utf-8") as src_f, \
                open(os.path.join(data_dir, f"data_ps.descriptions.{split}"), mode="r", encoding="utf-8") as tgt_f, \
                open(os.path.join(data_dir, f"data_ps.metadata.{split}"), mode="r", encoding="utf-8") as meta_f:
            sources = src_f.readlines()
            targets = tgt_f.readlines()
            metas = meta_f.readlines()
        assert len(sources) == len(targets) == len(metas)
        for source, target, meta in tqdm(zip(sources, targets, metas), desc="Reading", total=len(sources)):
            source = escape_code(source)
            target = escape_nl(target)
            idx = "#L".join(meta.strip().split())
            instances.append(
                DataInstance(
                    inputs=source.strip(),
                    outputs=target.strip(),
                    split=split,
                    idx=idx
                )
            )
            sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_code_docstring_corpus_gen(data_dir):
    def escape_code(s):
        s = s.strip()
        # s = s.replace(" DCNL DCSP ", "\n")
        s = s.replace(" DCNL DCSP ", "\n\t")
        s = s.replace(" DCNL  DCSP ", "\n\t")
        s = s.replace(" DCNL ", "\n")
        s = s.replace(" DCSP ", " ")
        while "\t " in s:
            s = s.replace("\t ", "\t\t")
        return s

    def escape_body(s):
        s = escape_code(s)
        return s.replace("DCSP ", "\t")

    def escape_nl(s):
        s = s.strip("'")
        s = s.strip()
        if "DCNL" in s:
            s = s.split("DCNL")[0]
        s = s.replace("DCSP", " ")
        return " ".join(s.split())

    data_dir = os.path.join(data_dir, "code_docstring_corpus")

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }

    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"data_ps.declarations.{split}"), mode="r", encoding="utf-8") as decl_f, \
                open(os.path.join(data_dir, f"data_ps.descriptions.{split}"), mode="r", encoding="utf-8") as desc_f, \
                open(os.path.join(data_dir, f"data_ps.bodies.{split}"), mode="r", encoding="utf-8") as body_f, \
                open(os.path.join(data_dir, f"data_ps.metadata.{split}"), mode="r", encoding="utf-8") as meta_f:
            decls = decl_f.readlines()
            descs = desc_f.readlines()
            targets = body_f.readlines()
            metas = meta_f.readlines()
        assert len(decls) == len(descs) == len(targets) == len(metas)

        for decl, desc, target, meta in zip(decls, descs, targets, metas):
            decl = escape_code(decl)
            desc = escape_nl(desc)
            target = escape_body(target)
            idx = "#L".join(meta.strip().split())
            instances.append(
                DataInstance(
                    inputs=[decl.strip(), desc.strip()],
                    outputs=target.strip(),
                    split=split,
                    idx=idx
                )
            )
            sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_concode(data_dir):
    data_dir = os.path.join(data_dir, "concode")

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"{split}.jsonl"), mode="r", encoding="utf-8") as f:
            lines = f.readlines()
        for idx, line in enumerate(tqdm(lines, total=len(lines), desc=f"Loading {split} data")):
            js = json.loads(line.strip())
            source = " ".join(js["nl"])
            target = " ".join(js["code"])
            instances.append(
                DataInstance(
                    inputs=source,
                    outputs=target,
                    split=split,
                    idx=str(idx)
                )
            )
            sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_many_types_4_typescript(data_dir):
    def convert_target(tokens: list[str], labels: list[str]):
        assert len(tokens) == len(labels)
        results = [f"{token}: {label}" for token, label in zip(tokens, labels) if label is not None and label != "null"]
        if len(results) == 0:
            results.append("None")
        return results

    data_dir = os.path.join(data_dir, "many_types_4_typescript")
    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"{split}.jsonl"), mode="r", encoding="utf-8") as f:
            lines = f.readlines()
        for idx, line in enumerate(tqdm(lines, total=len(lines), desc=f"Loading {split} data")):
            js = json.loads(line.strip())
            source = " ".join(js["tokens"])
            target = convert_target(tokens=js["tokens"], labels=js["labels"])
            idx = "#".join([js["url"], js["path"], js["commit_hash"]])
            instances.append(
                DataInstance(
                    inputs=source,
                    outputs=target,
                    split=split,
                    idx=str(idx)
                )
            )
            sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_codexglue_method_generation(data_dir):
    data_dir = os.path.join(data_dir, "codexglue_method_generation")

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"{split}.jsonl"), mode="r", encoding="utf-8") as f:
            lines = f.readlines()
        for idx, line in enumerate(tqdm(lines, total=len(lines), desc=f"Loading {split} data")):
            js = json.loads(line.strip())
            sign = js["signature"]
            nl = js["docstring"]
            target = js["body"]
            instances.append(
                DataInstance(
                    inputs=[sign, nl],
                    outputs=target,
                    split=split,
                    idx=js["id"]
                )
            )
            sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_hybrid_deep_com(data_dir):
    data_dir = os.path.join(data_dir, "hybrid_deep_com")

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"{split}.token.code"), mode="r", encoding="utf-8") as src_f, \
                open(os.path.join(data_dir, f"{split}.token.nl"), mode="r", encoding="utf-8") as tgt_f:
            sources = src_f.readlines()
            targets = tgt_f.readlines()
        assert len(sources) == len(targets)
        for idx, (source, target) in enumerate(tqdm(zip(sources, targets), desc="Reading", total=len(sources))):
            source = " ".join(source.strip().split())
            target = " ".join(target.strip().split())
            instances.append(
                DataInstance(
                    inputs=source,
                    outputs=target,
                    split=split,
                    idx=str(idx)
                )
            )
            sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_codenn(data_dir, subset):
    assert subset in ["csharp", "python", "sql"]
    data_dir = os.path.join(data_dir, "codenn", subset)

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"{split}.txt"), mode="r", encoding="utf-8") as f:
            lines = f.readlines()
        for idx, line in enumerate(tqdm(lines, total=len(lines), desc=f"Loading {split} data")):
            _, idx, nl, code, _ = line.strip().split("\t")
            instances.append(
                DataInstance(
                    inputs=code,
                    outputs=nl,
                    split=split,
                    idx=idx
                )
            )
            sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_code_qa(data_dir, subset):
    assert subset in ["java", "python"]
    data_dir = os.path.join(data_dir, "code_qa", subset)

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"{split}.code.original"), mode="r", encoding="utf-8") as code_f, \
                open(os.path.join(data_dir, f"{split}.question"), mode="r", encoding="utf-8") as q_f, \
                open(os.path.join(data_dir, f"{split}.answer"), mode="r", encoding="utf-8") as a_f:
            codes = code_f.readlines()
            questions = q_f.readlines()
            answers = a_f.readlines()
        assert len(codes) == len(questions) == len(answers)
        for idx, (code, q, a) in enumerate(tqdm(zip(codes, questions, answers), desc="Reading", total=len(codes))):
            instances.append(
                DataInstance(
                    inputs=[code.strip(), q.strip()],
                    outputs=a.strip(),
                    split=split,
                    idx=str(idx)
                )
            )
            sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_d2a(data_dir):
    import csv
    data_dir = os.path.join(data_dir, "d2a")
    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
    }
    for split in ["train", "valid"]:
        with open(os.path.join(data_dir, f"{split}.csv"), mode="r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                instances.append(
                    DataInstance(
                        inputs=row["code"].strip(),
                        outputs="Yes" if row["label"] == "1" else "No",
                        split=split,
                        idx=row["id"]
                    )
                )
                sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_code_gadget_database(data_dir, subset):
    assert subset in ["buffer", "resource"]
    data_dir = os.path.join(data_dir, "code_gadget_database")
    instances = []
    sizes = {
        "total": 0
    }
    with open(os.path.join(data_dir, "cwe{}_cgd.txt".format("119" if subset == "buffer" else "399")), mode="r",
              encoding="utf-8") as f:
        examples = f.read().strip().split("---------------------------------")
        for example in examples:
            lines = example.strip().split("\n")
            if len(lines) < 3:
                continue
            title = "#".join(lines[0].strip().split())
            label = "Yes" if lines[-1].strip() == "1" else "No"
            code = "\n".join(lines[1:-1]).strip()
            instances.append(
                DataInstance(
                    inputs=code,
                    outputs=label,
                    split="",
                    idx=title
                )
            )
            sizes["total"] += 1
    return instances, sizes


def read_sevc(data_dir, subset):
    assert subset in ["api", "arithmetic", "array", "pointer"]
    data_dir = os.path.join(data_dir, "sevc")
    instances = []
    sizes = {
        "total": 0
    }
    with open(os.path.join(data_dir, f"{subset}.txt"), mode="r", encoding="utf-8") as f:
        examples = f.read().strip().split("------------------------------")
        for example in examples:
            lines = example.strip().split("\n")
            if len(lines) < 3:
                continue
            title = "#".join(lines[0].strip().split())
            label = "Yes" if lines[-1].strip() == "1" else "No"
            code = "\n".join(lines[1:-1]).strip()
            instances.append(
                DataInstance(
                    inputs=code,
                    outputs=label,
                    split="",
                    idx=title
                )
            )
            sizes["total"] += 1
    return instances, sizes


def read_draper(data_dir):
    import pandas as pd
    import h5py
    data_dir = os.path.join(data_dir, "draper")
    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    for split in ["train", "valid", "test"]:
        with h5py.File(os.path.join(data_dir, f"VDISC_{split}.hdf5"), "r") as data:
            mydf = pd.DataFrame(list(data['functionSource']))
            mydf['CWE-119'] = list(data['CWE-119'])
            mydf['CWE-120'] = list(data['CWE-120'])
            mydf['CWE-469'] = list(data['CWE-469'])
            mydf['CWE-476'] = list(data['CWE-476'])
            mydf['CWE-other'] = list(data['CWE-other'])
        mydf.rename(columns={0: 'functionSource'}, inplace=True)
        for col in range(1, 6):
            mydf.iloc[:, col] = mydf.iloc[:, col].map({False: 0, True: 1})
        codes = [s.decode() for s in list(mydf["functionSource"])]
        labels = []
        for a, b, c, d, e in zip(mydf["CWE-119"], mydf["CWE-120"], mydf["CWE-469"], mydf["CWE-476"], mydf["CWE-other"]):
            if a + b + c + d + e > 0:
                labels.append(1)
            else:
                labels.append(0)
        assert len(codes) == len(labels)
        for idx, (code, label) in enumerate(zip(codes, labels)):
            instances.append(
                DataInstance(
                    inputs=code,
                    outputs="Yes" if label == 1 else "No",
                    split=split,
                    idx=str(idx),
                )
            )
            sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_buffer_overrun(data_dir):
    data_dir = os.path.join(data_dir, "buffer_overrun")
    instances = []
    sizes = {
        "train": 0,
        "test": 0
    }
    global_idx = 0
    with open(os.path.join(data_dir, "training_100.txt"), mode="r", encoding="utf-8") as src_f, \
            open(os.path.join(data_dir, "training_100_labels.txt"), mode="r", encoding="utf-8") as tgt_f:
        codes = []
        labels = []
        start = 0
        src_lines = src_f.read().split("\n")
        for target in tgt_f.readlines():
            end, label = target.strip().split(":=:")
            end = int(end)
            codes.append("\n".join(src_lines[start: end + 1]))
            if label == "0":
                labels.append("Yes")
            else:
                labels.append("No")
            start = end + 1
        assert len(codes) == len(labels)
        for code, label in zip(codes, labels):
            instances.append(
                DataInstance(
                    inputs=code.strip(),
                    outputs=label.strip(),
                    split="train",
                    idx=str(global_idx),
                )
            )
            global_idx += 1
            sizes["train"] += 1

    for name in ["1", "2", "3", "4"]:
        with open(os.path.join(data_dir, f"test_{name}_100.txt"), mode="r", encoding="utf-8") as src_f, \
                open(os.path.join(data_dir, f"test_{name}_100_labels.txt"), mode="r", encoding="utf-8") as tgt_f:
            codes = []
            labels = []
            start = 0
            src_lines = src_f.read().split("\n")
            for target in tgt_f.readlines():
                end, label = target.strip().split(":=:")
                end = int(end)
                codes.append("\n".join(src_lines[start: end + 1]))
                if label == "0":
                    labels.append("Yes")
                else:
                    labels.append("No")
                start = end + 1
            assert len(codes) == len(labels)
            for code, label in zip(codes, labels):
                instances.append(
                    DataInstance(
                        inputs=code.strip(),
                        outputs=label.strip(),
                        split="test",
                        idx=str(global_idx),
                    )
                )
                global_idx += 1
                sizes["test"] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_conala(data_dir, mode):
    assert mode in ["sum", "gen"]
    data_dir = os.path.join(data_dir, "conala")
    instances = []
    sizes = {
        "train": 0,
        "test": 0,
        "mined": 0
    }
    q_id_to_rewritten_intent = {}
    for split in ["train", "test"]:
        with open(os.path.join(data_dir, f"conala-{split}.json"), mode="r", encoding="utf-8") as f:
            data = json.load(f)
        for item in data:
            q_id = item["question_id"]
            code = item["snippet"].strip()
            if item["rewritten_intent"] == None:
                if q_id in q_id_to_rewritten_intent:
                    nl = q_id_to_rewritten_intent[q_id]
                else:
                    nl = item["intent"]
            else:
                nl = item["rewritten_intent"].strip()
                q_id_to_rewritten_intent[q_id] = nl
            instances.append(
                DataInstance(
                    inputs=code if mode == "sum" else nl,
                    outputs=nl if mode == "sum" else code,
                    split=split,
                    idx=f"{q_id}"
                )
            )
            sizes[split] += 1
    with open(os.path.join(data_dir, f"conala-mined.jsonl"), mode="r", encoding="utf-8") as f:
        lines = f.readlines()
    for line in lines:
        js = json.loads(line.strip())
        code = js["snippet"].strip()
        nl = js["intent"].strip()
        instances.append(
            DataInstance(
                inputs=code if mode == "sum" else nl,
                outputs=nl if mode == "sum" else code,
                split="mined",
                idx=js["id"]
            )
        )
        sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_xlcost_translate(data_dir, source_lang, target_lang, mode):
    assert source_lang in ["C++", "C", "Javascript", "PHP", "Python", "C#", "Java"]
    assert target_lang in ["C++", "C", "Javascript", "PHP", "Python", "C#", "Java"]
    assert mode in ["snippet", "program"]

    def parse_code(s):
        s = s.strip()
        s = s.replace("NEWLINE", "\n")
        s = s.replace("NEW_LINE", "\n")
        s = s.replace("INDENT", "\t")
        return s.strip()

    extension_map = {
        "C++": "cpp",
        "C": "c",
        "Javascript": "js",
        "PHP": "php",
        "Python": "py",
        "C#": "cs",
        "Java": "java"
    }
    source_ext = extension_map[source_lang]
    target_ext = extension_map[target_lang]

    data_dir = os.path.join(data_dir,
                            "xlcost",
                            "generation",
                            "pair_data_tok_{}".format("1" if mode == "snippet" else "full"))
    if os.path.exists(os.path.join(data_dir, f"{source_lang}-{target_lang}")):
        task_name = f"{source_lang}-{target_lang}"
    else:
        task_name = f"{target_lang}-{source_lang}"
    data_dir = os.path.join(data_dir, task_name)

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0,
    }

    for split, split_name in zip(["train", "valid", "test"], ["train", "val", "test"]):
        with open(os.path.join(data_dir, f"{split_name}-{task_name}-tok.{source_ext}"),
                  mode="r", encoding="utf-8") as src_f, \
                open(os.path.join(data_dir, f"{split_name}-{task_name}-tok.{target_ext}"),
                     mode="r", encoding="utf-8") as tgt_f, \
                open(os.path.join(data_dir, f"{split_name}-{source_lang}-map.jsonl"),
                     mode="r", encoding="utf-8") as src_idx_f, \
                open(os.path.join(data_dir, f"{split_name}-{target_lang}-map.jsonl"),
                     mode="r", encoding="utf-8") as tgt_idx_f:
            sources = src_f.readlines()
            targets = tgt_f.readlines()
            source_indices = src_idx_f.readlines()
            target_indices = tgt_idx_f.readlines()
        assert len(sources) == len(targets) == len(source_indices) == len(target_indices)
        for source, target, src_idx, tgt_idx in zip(sources, targets, source_indices, target_indices):
            source = parse_code(source)
            target = parse_code(target)
            instances.append(
                DataInstance(
                    inputs=source,
                    outputs=target,
                    split=split,
                    idx=f"{src_idx}#{tgt_idx}"
                )
            )
            sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_xlcost_summarization(data_dir, source_lang, mode):
    assert source_lang in ["C++", "C", "Javascript", "PHP", "Python", "C#", "Java"]
    assert mode in ["snippet", "program"]

    target_lang = "comment" if mode == "snippet" else "desc"

    def parse_code(s):
        s = s.strip()
        s = s.replace("NEWLINE", "\n")
        s = s.replace("NEW_LINE", "\n")
        s = s.replace("INDENT", "\t")
        return s.strip()

    extension_map = {
        "C++": "cpp",
        "C": "c",
        "Javascript": "js",
        "PHP": "php",
        "Python": "py",
        "C#": "cs",
        "Java": "java"
    }
    source_ext = extension_map[source_lang]
    target_ext = "txt"

    data_dir = os.path.join(data_dir,
                            "xlcost",
                            "generation",
                            "pair_data_tok_{}_{}".format("1" if mode == "snippet" else "full", target_lang))
    task_name = f"{source_lang}-{target_lang}"
    data_dir = os.path.join(data_dir, task_name)

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0,
    }

    for split, split_name in zip(["train", "valid", "test"], ["train", "val", "test"]):
        with open(os.path.join(data_dir, f"{split_name}-{task_name}-tok.{source_ext}"),
                  mode="r", encoding="utf-8") as src_f, \
                open(os.path.join(data_dir, f"{split_name}-{task_name}-tok.{target_ext}"),
                     mode="r", encoding="utf-8") as tgt_f, \
                open(os.path.join(data_dir, f"{split_name}-{source_lang}-map.jsonl"),
                     mode="r", encoding="utf-8") as src_idx_f, \
                open(os.path.join(data_dir, f"{split_name}-{target_lang}-map.jsonl"),
                     mode="r", encoding="utf-8") as tgt_idx_f:
            sources = src_f.readlines()
            targets = tgt_f.readlines()
            source_indices = src_idx_f.readlines()
            target_indices = tgt_idx_f.readlines()
        assert len(sources) == len(targets) == len(source_indices) == len(target_indices)
        for source, target, src_idx, tgt_idx in zip(sources, targets, source_indices, target_indices):
            assert src_idx == tgt_idx
            source = parse_code(source)
            target = target.strip()
            instances.append(
                DataInstance(
                    inputs=source,
                    outputs=target,
                    split=split,
                    idx=f"{src_idx}"
                )
            )
            sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_xlcost_gen(data_dir, source_lang, mode):
    assert source_lang in ["C++", "C", "Javascript", "PHP", "Python", "C#", "Java"]
    assert mode in ["snippet", "program"]

    target_lang = "comment" if mode == "snippet" else "desc"

    def parse_code(s):
        s = s.strip()
        s = s.replace("NEWLINE", "\n")
        s = s.replace("NEW_LINE", "\n")
        s = s.replace("INDENT", "\t")
        return s.strip()

    extension_map = {
        "C++": "cpp",
        "C": "c",
        "Javascript": "js",
        "PHP": "php",
        "Python": "py",
        "C#": "cs",
        "Java": "java"
    }
    source_ext = extension_map[source_lang]
    target_ext = "txt"

    data_dir = os.path.join(data_dir,
                            "xlcost",
                            "generation",
                            "pair_data_tok_{}".format("1_comment" if mode == "snippet" else "full_desc_comment"))
    task_name = f"{source_lang}-{target_lang}"
    data_dir = os.path.join(data_dir, task_name)

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0,
    }

    for split, split_name in zip(["train", "valid", "test"], ["train", "val", "test"]):
        with open(os.path.join(data_dir, f"{split_name}-{task_name}-tok.{source_ext}"),
                  mode="r", encoding="utf-8") as src_f, \
                open(os.path.join(data_dir, f"{split_name}-{task_name}-tok.{target_ext}"),
                     mode="r", encoding="utf-8") as tgt_f, \
                open(os.path.join(data_dir, f"{split_name}-{source_lang}-map.jsonl"),
                     mode="r", encoding="utf-8") as src_idx_f, \
                open(os.path.join(data_dir, f"{split_name}-{target_lang}-map.jsonl"),
                     mode="r", encoding="utf-8") as tgt_idx_f:
            sources = src_f.readlines()
            targets = tgt_f.readlines()
            source_indices = src_idx_f.readlines()
            target_indices = tgt_idx_f.readlines()
        assert len(sources) == len(targets) == len(source_indices) == len(target_indices)
        for source, target, src_idx, tgt_idx in zip(sources, targets, source_indices, target_indices):
            assert src_idx == tgt_idx
            source = parse_code(source)
            target = target.strip()
            instances.append(
                DataInstance(
                    inputs=target,
                    outputs=source,
                    split=split,
                    idx=f"{src_idx}"
                )
            )
            sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_avatar(data_dir, source_lang, target_lang):
    assert source_lang in ["java", "python"]
    assert target_lang in ["java", "python"]
    data_dir = os.path.join(data_dir, "avatar")
    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"{split}.java-python.{source_lang}"), mode="r", encoding="utf-8") as src_f, \
                open(os.path.join(data_dir, f"{split}.java-python.{target_lang}"), mode="r", encoding="utf-8") as tgt_f, \
                open(os.path.join(data_dir, f"{split}.java-python.id"), mode="r", encoding="utf-8") as idx_f:
            sources = src_f.readlines()
            targets = tgt_f.readlines()
            indices = idx_f.readlines()
        assert len(sources) == len(targets) == len(indices)
        for source, target, idx in zip(sources, targets, indices):
            instances.append(
                DataInstance(
                    inputs=source.strip(),
                    outputs=target.strip(),
                    split=split,
                    idx=idx.strip()
                )
            )
            sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_many_types_4_py(data_dir):
    def convert_target(tokens: str, types: str):
        tokens = tokens.strip().split()
        types = types.strip().split()
        if len(tokens) != len(types):
            return None
        results = [f"{token}: {label}" for token, label in zip(tokens, types) if label != "0"]
        if len(results) == 0:
            results.append("None")
        return results

    data_dir = os.path.join(data_dir, "many_types_4_py")

    file_to_split = {}
    with open(os.path.join(data_dir, f"dataset_split.csv"), mode="r", encoding="utf-8") as f:
        for line in f.readlines():
            split, file = line.strip().split(",", 1)
            split = split.strip()
            file = file.strip().strip("\"")
            assert split in ["train", "valid", "test"]
            file_to_split[file] = split

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    data_dir = os.path.join(data_dir, "processed_projects_clean")
    for filename in os.listdir(data_dir):
        file_path = os.path.join(data_dir, filename)
        if filename.endswith(".json") and os.path.isfile(file_path):
            with open(file_path, mode="r", encoding="utf-8") as f:
                js = json.load(f)
                for _, repo_data in js.items():
                    src_files_dict = repo_data["src_files"]
                    for file, data in src_files_dict.items():
                        source = data["untyped_seq"]
                        if len(source) == 0:
                            continue
                        target = convert_target(source, data["typed_seq"])
                        if target is None or (len(target) == 1 and target[0] == "None"):
                            continue
                        source = source.replace("[EOL]", "\n")
                        if file == "repos/yannhyu/breadnbutter/SOAP_WS/suds-jurko-0.6/\ntest_with_sample_cases.py":
                            split = "train"
                        else:
                            split = file_to_split[file]
                        instances.append(
                            DataInstance(
                                inputs=source,
                                outputs=target,
                                split=split,
                                idx=file
                            )
                        )
                        sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_commit_gen(data_dir):
    data_dir = os.path.join(data_dir, "commit_gen")

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    global_idx = 0
    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"cleaned.{split}.diff"), mode="r", encoding="utf-8") as src_f, \
                open(os.path.join(data_dir, f"cleaned.{split}.msg"), mode="r", encoding="utf-8") as tgt_f:
            sources = src_f.readlines()
            targets = tgt_f.readlines()
        assert len(sources) == len(targets)
        for source, target in tqdm(zip(sources, targets), desc="Reading", total=len(sources)):
            instances.append(
                DataInstance(
                    inputs=source.strip(),
                    outputs=target.strip(),
                    split=split,
                    idx=str(global_idx)
                )
            )
            sizes[split] += 1
            global_idx += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_code_comment_cls(data_dir):
    from openpyxl import load_workbook

    data_dir = os.path.join(data_dir, "comment_cls")

    instances = []
    sizes = {
        "total": 0
    }

    wb = load_workbook(os.path.join(data_dir, "data_set.xlsx"))
    sheet = wb.worksheets[0]
    rows = list(sheet.rows)
    for idx, row in enumerate(rows[1:]):
        code = row[0].value
        nl = row[1].value
        label = row[2].value
        if code is None or nl is None or label is None:
            continue
        instances.append(
            DataInstance(
                inputs=[code.strip(), nl.strip()],
                outputs=label.strip(),
                split="",
                idx=str(idx)
            )
        )
        sizes["total"] += 1
    return instances, sizes


def read_comment_cls(data_dir):
    from openpyxl import load_workbook

    data_dir = os.path.join(data_dir, "comment_cls")

    instances = []
    sizes = {
        "total": 0
    }

    wb = load_workbook(os.path.join(data_dir, "data_set.xlsx"))
    sheet = wb.worksheets[0]
    rows = list(sheet.rows)
    for idx, row in enumerate(rows[1:]):
        nl = row[1].value
        label = row[2].value
        if nl is None or label is None:
            continue
        instances.append(
            DataInstance(
                inputs=nl.strip(),
                outputs=label.strip(),
                split="",
                idx=str(idx)
            )
        )
        sizes["total"] += 1
    return instances, sizes


def read_atom(data_dir):
    import csv
    csv.field_size_limit(500 * 1024 * 1024)

    data_dir = os.path.join(data_dir, "atom")
    instances = []
    sizes = {
        "total": 0
    }
    for filename in os.listdir(data_dir):
        file_path = os.path.join(data_dir, filename)
        if filename.endswith(".csv") and os.path.isfile(file_path):
            print(file_path)
            with open(file_path, mode="r", encoding="utf-8") as f:
                reader = csv.reader((line.replace('\0', '') for line in f), delimiter=",")
                header = next(reader)
                for row in reader:
                    assert len(row) == 10
                    diff = row[3]
                    commit = row[1]
                    project = row[9]
                    commit_id = row[0]
                    instances.append(
                        DataInstance(
                            inputs=diff.strip(),
                            outputs=commit.strip(),
                            split="",
                            idx=f"{project}#{commit_id}"
                        )
                    )
                    sizes["total"] += 1
    return instances, sizes


def read_pseudo_gen(data_dir, gen_type):
    assert gen_type in ["code", "pseudo"]
    source_ext = "anno" if gen_type == "code" else "code"
    target_ext = "code" if gen_type == "code" else "anno"

    data_dir = os.path.join(data_dir, "pseudo_gen")

    instances = []
    sizes = {
        "total": 0
    }
    with open(os.path.join(data_dir, f"all.{source_ext}"), mode="r", encoding="utf-8") as src_f, \
            open(os.path.join(data_dir, f"all.{target_ext}"), mode="r", encoding="utf-8") as tgt_f:
        sources = src_f.readlines()
        targets = tgt_f.readlines()
    assert len(sources) == len(targets)
    for idx, (source, target) in enumerate(tqdm(zip(sources, targets), desc="Reading", total=len(sources))):
        instances.append(
            DataInstance(
                inputs=source.strip(),
                outputs=target.strip(),
                split="",
                idx=str(idx)
            )
        )
        sizes["total"] += 1
    return instances, sizes


def read_java_large_method_name(data_dir):
    data_dir = os.path.join(data_dir, "java_large")

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"{split}.jsonl"), mode="r", encoding="utf-8") as f:
            lines = f.readlines()
        for idx, line in enumerate(tqdm(lines, desc="Reading", total=len(lines))):
            js = json.loads(line.strip())
            code = js["code"].strip()
            name = js["name"].strip()
            instances.append(
                DataInstance(
                    inputs=code,
                    outputs=name,
                    split=split,
                    idx=str(idx)
                )
            )
            sizes[split] += 1
    sizes["total"] = len(instances)
    return instances, sizes


def read_so_ds(data_dir):
    data_dir = os.path.join(data_dir, "so_ds")

    instances = []
    sizes = {
        "total": 0
    }
    with open(os.path.join(data_dir, "so-ds-feb20.jsonl"), mode="r", encoding="utf-8") as f:
        lines = f.readlines()
    for idx, line in enumerate(tqdm(lines, desc="Reading", total=len(lines))):
        js = json.loads(line.strip())
        code = js["code"].strip()
        desc = js["description"].strip()
        idx = js["id"].strip()
        instances.append(
            DataInstance(
                inputs=desc,
                outputs=code,
                split="",
                idx=idx,
            )
        )
        sizes["total"] += 1
    assert sizes["total"] == len(instances)
    return instances, sizes


def read_ncs(data_dir):
    data_dir = os.path.join(data_dir, "ncs")

    instances = []
    sizes = {
        "total": 0
    }
    with open(os.path.join(data_dir, "287_android_questions.json"), mode="r", encoding="utf-8") as f:
        data = json.load(f)
    for js in data:
        desc = js["question"].strip()
        code = js["answer"].strip()
        idx = js["stackoverflow_id"].strip()
        instances.append(
            DataInstance(
                inputs=desc,
                outputs=code,
                split="",
                idx=idx,
            )
        )
        sizes["total"] += 1
    assert sizes["total"] == len(instances)
    return instances, sizes


def read_nl_2_bash(data_dir, mode):
    assert mode in ["gen", "sum"]
    if mode == "gen":
        src_name = "nl"
        tgt_name = "cm"
    else:
        src_name = "cm"
        tgt_name = "nl"
    data_dir = os.path.join(data_dir, "nl_2_bash")

    instances = []
    sizes = {
        "total": 0
    }
    with open(os.path.join(data_dir, f"all.{src_name}.filtered"), mode="r", encoding="utf-8") as src_f, \
            open(os.path.join(data_dir, f"all.{tgt_name}.filtered"), mode="r", encoding="utf-8") as tgt_f:
        sources = src_f.readlines()
        targets = tgt_f.readlines()
    assert len(sources) == len(targets)
    for idx, (source, target) in enumerate(tqdm(zip(sources, targets), desc="Reading", total=len(sources))):
        source = source.strip()
        target = target.strip()
        instances.append(
            DataInstance(
                inputs=source.strip(),
                outputs=target.strip(),
                split="",
                idx=str(idx)
            )
        )
        sizes["total"] += 1
    return instances, sizes


def read_kb13(data_dir):
    import csv
    data_dir = os.path.join(data_dir, "kb13")
    instances = []
    sizes = {
        "total": 0,
    }
    with open(os.path.join(data_dir, f"regexp-naacl2013-data.csv"), mode="r", encoding="utf-8") as f:
        reader = csv.reader(f)
        idx = 0
        for row in reader:
            instances.append(
                DataInstance(
                    inputs=row[0].strip(),
                    outputs=row[1].strip(),
                    split="",
                    idx=str(idx)
                )
            )
            idx += 1
            sizes["total"] += 1
    return instances, sizes


def read_nl_rx(data_dir):
    data_dir = os.path.join(data_dir, "nl_rx")

    instances = []
    sizes = {
        "total": 0
    }
    with open(os.path.join(data_dir, "src.txt"), mode="r", encoding="utf-8") as src_f, \
            open(os.path.join(data_dir, "targ.txt"), mode="r", encoding="utf-8") as tgt_f:
        sources = src_f.readlines()
        targets = tgt_f.readlines()
    assert len(sources) == len(targets)
    for idx, (source, target) in enumerate(tqdm(zip(sources, targets), desc="Reading", total=len(sources))):
        source = source.strip()
        target = target.strip()
        instances.append(
            DataInstance(
                inputs=source.strip(),
                outputs=target.strip(),
                split="",
                idx=str(idx)
            )
        )
        sizes["total"] += 1
    return instances, sizes


def read_spoc(data_dir):
    import csv
    data_dir = os.path.join(data_dir, "spoc")
    instances = []
    sizes = {
        "train-eval": 0,
        "train-train": 0,
        "train-test": 0,
        "testp": 0,
        "testw": 0
    }
    for split in sizes.keys():
        with open(os.path.join(data_dir, f"spoc-{split}.tsv"), mode="r", encoding="utf-8") as f:
            reader = csv.reader(f, delimiter="\t")
            header = next(reader)
            for row in reader:
                pseudo = row[0].strip()
                code = row[1].strip()
                idx = "#".join(row[2:])
                if pseudo == "" or code == "":
                    continue
                instances.append(
                    DataInstance(
                        inputs=pseudo,
                        outputs=code,
                        split=split,
                        idx=idx
                    )
                )
                sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_algo_lisp(data_dir):
    import csv

    data_dir = os.path.join(data_dir, "algo_lisp")
    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    global_idx = 0
    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"{split}.tsv"), mode="r", encoding="utf-8") as f:
            reader = csv.reader(f, delimiter="\t")
            for row in reader:
                nl = row[0].strip()
                code = row[1].strip()
                if nl == "" or code == "":
                    continue
                instances.append(
                    DataInstance(
                        inputs=nl,
                        outputs=code,
                        split=split,
                        idx=str(global_idx)
                    )
                )
                sizes[split] += 1
                global_idx += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_deep_api(data_dir):
    import csv

    data_dir = os.path.join(data_dir, "deep_api")
    instances = []
    sizes = {
        "train": 0,
        "test": 0
    }
    global_idx = 0
    for split in ["train", "test"]:
        with open(os.path.join(data_dir, f"{split}.tsv"), mode="r", encoding="utf-8") as f:
            reader = csv.reader(f, delimiter="\t")
            for row in reader:
                nl = row[0].strip()
                api = " ".join(row[1].strip().split())
                if nl == "" or api == "":
                    continue
                instances.append(
                    DataInstance(
                        inputs=nl,
                        outputs=api,
                        split=split,
                        idx=str(global_idx)
                    )
                )
                sizes[split] += 1
                global_idx += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_code_search_net_method_name(data_dir, subset):
    assert subset in ["java", "python", "javascript", "php", "go", "ruby"]

    data_dir = os.path.join(data_dir, "code_search_net_filtered", subset)

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"{split}.jsonl"), mode="r", encoding="utf-8") as f:
            lines = f.readlines()
        for idx, line in enumerate(tqdm(lines, total=len(lines), desc=f"Loading {split} data")):
            js = json.loads(line.strip())
            source = " ".join(js["code_tokens"])
            target = js["func_name"].strip()
            source = source.replace(target, "f")
            instances.append(
                DataInstance(
                    inputs=source,
                    outputs=target,
                    split=split,
                    idx=js["path"]
                )
            )
            sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_naps(data_dir):
    data_dir = os.path.join(data_dir, "naps")

    instances = []
    sizes = {
        "trainB": 0,
        "test": 0
    }
    for split in sizes.keys():
        with open(os.path.join(data_dir, f"naps.{split}.1.0.jsonl"), mode="r", encoding="utf-8") as f:
            lines = f.readlines()
        for line in tqdm(lines, total=len(lines), desc=f"Loading {split} data"):
            js = json.loads(line.strip())
            code = " ".join(js["code_sequence"])
            nl = " ".join(js["text"])
            idx = js["entry_id"]
            instances.append(
                DataInstance(
                    inputs=nl,
                    outputs=code,
                    split=split,
                    idx=idx
                )
            )
            sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_shell_code_ia32(data_dir):
    import csv

    data_dir = os.path.join(data_dir, "shell_code_ia32")
    instances = []
    sizes = {
        "total": 0
    }
    global_idx = 0
    with open(os.path.join(data_dir, "Shellcode_IA32.tsv"), mode="r", encoding="utf-8") as f:
        reader = csv.reader(f, delimiter="\t")
        header = next(reader)
        for row in reader:
            shell = row[0].strip()
            intent = row[1].strip()
            if shell == "" or intent == "":
                continue
            instances.append(
                DataInstance(
                    inputs=intent,
                    outputs=shell,
                    split="",
                    idx=str(global_idx)
                )
            )
            sizes["total"] += 1
            global_idx += 1
    return instances, sizes


def read_evil(data_dir, subset):
    assert subset in ["encoder", "decoder"]
    data_dir = os.path.join(data_dir, "evil", subset)
    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    global_idx = 0
    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"{subset}-{split}.in"), mode="r", encoding="utf-8") as src_f, \
                open(os.path.join(data_dir, f"{subset}-{split}.out"), mode="r", encoding="utf-8") as tgt_f:
            sources = src_f.readlines()
            targets = tgt_f.readlines()
        assert len(sources) == len(targets)
        for source, target in tqdm(zip(sources, targets), desc="Reading", total=len(sources)):
            instances.append(
                DataInstance(
                    inputs=source.strip(),
                    outputs=target.strip(),
                    split=split,
                    idx=str(global_idx)
                )
            )
            sizes[split] += 1
            global_idx += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_t_fix(data_dir):
    data_dir = os.path.join(data_dir, "t_fix")
    instances = []
    sizes = {
        "total": 0
    }
    for split in ["eslint", "repo_specific"]:
        with open(os.path.join(data_dir, f"data_autofix_tracking_{split}_final.json"), mode="r", encoding="utf-8") as f:
            data = json.load(f)
        for item in tqdm(data):
            source = item["source_code"].strip()
            target = item["target_code"].strip()
            idx = "#".join([item["repo"], item["source_changeid"], item["target_changeid"]])
            instances.append(
                DataInstance(
                    inputs=source,
                    outputs=target,
                    split=split,
                    idx=idx
                )
            )
            sizes["total"] += 1
    return instances, sizes


def read_many_sstubs_4_j_sstubs_type(data_dir):
    data_dir = os.path.join(data_dir, "many_sstubs_4_j")
    instances = []
    sizes = {
        "total": 0
    }
    with open(os.path.join(data_dir, f"sstubsLarge.json"), mode="r", encoding="utf-8") as f:
        data = json.load(f)
    for item in tqdm(data):
        source = item["sourceBeforeFix"].strip()
        fixed = item["sourceAfterFix"].strip()
        target = item["bugType"].strip()
        target = " ".join(target.split("_")).title()
        idx = "#".join(
            [item["projectName"], item["commitSHA1"] if "commitSHA1" in item else item["fixCommitParentSHA1"]])
        instances.append(
            DataInstance(
                inputs=[source, fixed],
                outputs=target,
                split="",
                idx=idx
            )
        )
        sizes["total"] += 1
    return instances, sizes


def read_cod_rep(data_dir):
    data_dir = os.path.join(data_dir, "cod_rep")
    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    global_idx = 0
    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"src-{split}.txt"), mode="r", encoding="utf-8") as src_f, \
                open(os.path.join(data_dir, f"tgt-{split}.txt"), mode="r", encoding="utf-8") as tgt_f:
            sources = src_f.readlines()
            targets = tgt_f.readlines()
        assert len(sources) == len(targets)
        for source, target in tqdm(zip(sources, targets), desc="Reading", total=len(sources)):
            instances.append(
                DataInstance(
                    inputs=source.strip(),
                    outputs=target.strip(),
                    split=split,
                    idx=str(global_idx)
                )
            )
            sizes[split] += 1
            global_idx += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_apps(data_dir):
    from datasets import load_from_disk
    data_dir = os.path.join(data_dir, "apps")
    instances = []
    sizes = {
        "train": 0,
        "test": 0
    }
    dataset = load_from_disk(data_dir)
    global_idx = 0
    for split in sizes.keys():
        subset = dataset[split]
        for item in subset:
            instances.append(
                DataInstance(
                    inputs=item["question"].strip().strip("-"),
                    outputs=item["solutions"].strip().strip("[").strip("]").strip("\""),
                    split=split,
                    idx=str(global_idx)
                )
            )
            sizes[split] += 1
            global_idx += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_py_torrent(data_dir, mode):
    assert mode in ["sum", "gen"]
    data_dir = os.path.join(data_dir, "py_torrent")

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    for split in sizes.keys():
        subset_dir = os.path.join(data_dir, split)
        print(f"Split: {split}")
        for filename in os.listdir(subset_dir):
            file_path = os.path.join(subset_dir, filename)
            if os.path.isfile(file_path) and filename.endswith("jsonl"):
                with open(file_path, mode="r", encoding="utf-8") as f:
                    lines = f.readlines()
                for line in tqdm(lines, total=len(lines), desc=f"Loading {filename} data"):
                    js = json.loads(line.strip())
                    code = " ".join(js["code_tokens"])
                    nl = " ".join(js["docstring_tokens"])
                    idx = js["path"]
                    instances.append(
                        DataInstance(
                            inputs=code if mode == "sum" else nl,
                            outputs=nl if mode == "gen" else code,
                            split=split,
                            idx=idx
                        )
                    )
                    sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_juice(data_dir):
    data_dir = os.path.join(data_dir, "juice")

    instances = []
    sizes = {
        "valid": 0,
        "test": 0,
        "train": 0
    }
    for split in sizes.keys():
        print(f"Reading {split} data")
        with open(os.path.join(data_dir, f"{split}.jsonl"), mode="r", encoding="utf-8") as f:
            lines = f.readlines()
        for line in tqdm(lines, total=len(lines), desc=f"Loading {split} data"):
            js = json.loads(line.strip())
            code = " ".join(js["code_tokens_clean"])
            nl = " ".join(js["nl"])
            idx = "#".join([js["metadata"]["repo"], js["metadata"]["path"]])
            instances.append(
                DataInstance(
                    inputs=nl,
                    outputs=code,
                    split=split,
                    idx=idx
                )
            )
            sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_code_search_net_gen(data_dir, subset):
    assert subset in ["java", "python", "javascript", "php", "go", "ruby"]

    data_dir = os.path.join(data_dir, "code_search_net_filtered", subset)

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    for split in ["train", "valid", "test"]:
        with open(os.path.join(data_dir, f"{split}.jsonl"), mode="r", encoding="utf-8") as f:
            lines = f.readlines()
        for idx, line in enumerate(tqdm(lines, total=len(lines), desc=f"Loading {split} data")):
            js = json.loads(line.strip())
            source = " ".join(js["docstring_tokens"])
            target = " ".join(js["code_tokens"])
            instances.append(
                DataInstance(
                    inputs=source,
                    outputs=target,
                    split=split,
                    idx=js["path"]
                )
            )
            sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_fix_eval(data_dir, lang):
    assert lang in ["java", "python"]
    data_dir = os.path.join(data_dir, "fix_eval", lang, "processed")

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    for split in sizes.keys():
        with open(os.path.join(data_dir, f"src_{split}.{lang}-{lang}.{lang}"), mode="r", encoding="utf-8") as src_f, \
             open(os.path.join(data_dir, f"tgt_{split}.{lang}-{lang}.{lang}"), mode="r", encoding="utf-8") as tgt_f, \
             open(os.path.join(data_dir, f"{split}.{lang}-{lang}.id"), mode="r", encoding="utf-8") as idx_f:
            sources = src_f.readlines()
            targets = tgt_f.readlines()
            indices = idx_f.readlines()
        assert len(sources) == len(targets) == len(indices)
        for idx, source, target in tqdm(zip(indices, sources, targets), desc="Reading", total=len(sources)):
            instances.append(
                DataInstance(
                    inputs=source.strip(),
                    outputs=target.strip(),
                    split=split,
                    idx=idx.strip()
                )
            )
            sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes


def read_fix_eval_verdict(data_dir, lang):
    assert lang in ["java", "python"]
    data_dir = os.path.join(data_dir, "fix_eval", lang, "processed_with_verdict")

    instances = []
    sizes = {
        "train": 0,
        "valid": 0,
        "test": 0
    }
    for split in sizes.keys():
        with open(os.path.join(data_dir, f"src_{split}.{lang}-{lang}.{lang}"), mode="r", encoding="utf-8") as src_f, \
             open(os.path.join(data_dir, f"tgt_{split}.{lang}-{lang}.{lang}"), mode="r", encoding="utf-8") as tgt_f, \
             open(os.path.join(data_dir, f"{split}.{lang}-{lang}.id"), mode="r", encoding="utf-8") as idx_f:
            sources = src_f.readlines()
            targets = tgt_f.readlines()
            indices = idx_f.readlines()
        assert len(sources) == len(targets) == len(indices)
        for idx, source, target in tqdm(zip(indices, sources, targets), desc="Reading", total=len(sources)):
            source, verdict = source.strip().split("verdict:")
            instances.append(
                DataInstance(
                    inputs=[source.strip(), verdict.strip()],
                    outputs=target.strip(),
                    split=split,
                    idx=idx.strip()
                )
            )
            sizes[split] += 1
    assert sum(sizes.values()) == len(instances)
    sizes["total"] = len(instances)
    return instances, sizes
